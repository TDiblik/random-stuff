import random
import openpyxl
from faker import Faker
from faker_food import FoodProvider

fake = Faker('cs_CZ')
fake.add_provider(FoodProvider)

header = ["Jméno", "Oblíbené jídlo", "Počet snězení tohoto jídla"]
data = []
for _ in range(150):
    name = fake.unique.name()
    food = fake.dish()
    count = random.randint(1, 50)
    data.append([name, food, count])

workbook = openpyxl.Workbook()
worksheet = workbook.active
for col, header_value in enumerate(header, start=1):
    worksheet.cell(row=1, column=col, value=header_value)
for row, row_data in enumerate(data, start=2):
    for col, value in enumerate(row_data, start=1):
        worksheet.cell(row=row, column=col, value=value)

workbook.save(filename='file_to_process.xlsx')